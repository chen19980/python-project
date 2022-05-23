from bs4 import BeautifulSoup
import requests
import time
import random
import openpyxl

wb = openpyxl.Workbook() #開啟新的工作簿
ws = wb.active           #開啟新的工作表

res = requests.get("https://www.1111.com.tw/search/job?ks=%E5%A4%A7%E6%95%B8%E6%93%9A")

soup = BeautifulSoup(res.text)

page = 0
ws.append(["職位名稱","網址","公司名稱","公司地區","薪水",])   #放入list
ws["F1"] = "給薪方式"
ws["G1"] = "薪資下限"
ws["H1"] = "薪資上限"
while page<73:
    page += 1
    res = requests.get("https://www.1111.com.tw/search/job?ks=%E5%A4%A7%E6%95%B8%E6%93%9A&page="+str(page))
    soup = BeautifulSoup(res.text)
    for job in soup.find_all("div",class_="job-list-item")[0].find_all("div",class_="job_item_info") :
        # res2 = requests.get(job.a["href"])
        # soup2 = BeautifulSoup(res2.text)
        a = job.h5.text   #職位名稱
        b = job.a["href"]    #網址
        c = job.h6.text      #公司名稱
        d = job.find_all("a",class_="job_item_detail_location mr-3 position-relative")[0].text    #公司地區
        e = job.find_all("div", class_="job_item_detail_salary ml-3 font-weight-style digit_6")[0].text  #薪水
        f = job.select("div")[2].div.div.text[0:2]

        price = ""

        for word in e:
            if  word == "0" or word == "1" or word == "2" or word == "3" or word == "4" or word == "5" or word == "6" or word == "7" or word == "8" or word == "9" or word == "~" or word == "." or word == "萬" :
                price += word

        low_price = ""
        high_price = ""
        if "~" in price :
            low_price = price[0:price.find("~")]
            high_price = price[price.find("~")+1:len(price)]
        else:
            low_price = price
            high_price = price
 
        if "萬" in low_price:
            low_price = low_price[0:low_price.find("萬")]
            low_price = float(low_price)*10000 

        else:
            low_price = float(low_price)

 
        if "萬" in high_price:
            high_price = high_price[0:high_price.find("萬")]
            high_price = float(high_price)*10000 
        
        else:
            high_price = float(high_price)

        # print(soup2.find_all("div",class_="contacter_info")[0].p.text)  #聯絡人
        ws.append([a,b,c,d,e,f,low_price,high_price])
        wb.save("職缺清單20220406.xlsx")  #儲存工作簿
    print(f"已完成:{page}")   
    print("--------------") 
    time.sleep(2)












