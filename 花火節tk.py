import tkinter as tk
import openpyxl as op
import pymongo
import certifi

# wb = op.load_workbook("訂房網站.xlsx")
# ws = wb.active

client = pymongo.MongoClient("mongodb+srv://test:7ww17W3A5wBoWLkA@cluster0.jgt9o.mongodb.net/?retryWrites=true&w=majority", tlsCAFile=certifi.where())

db = client.trivago
col = db.roomlist

# for x in range(2,7812):
#     a = ws[f"A{x}"].value
#     b = ws[f"B{x}"].value
#     c = ws[f"C{x}"].value
#     d = ws[f"D{x}"].value
#     e = ws[f"E{x}"].value
#     f = ws[f"F{x}"].value
#     g = ws[f"G{x}"].value
#     h = ws[f"H{x}"].value
#     i = ws[f"I{x}"].value
#     j = ws[f"J{x}"].value
#     k = ws[f"K{x}"].value
#     l = ws[f"L{x}"].value
#     m = ws[f"M{x}"].value

#     wk = {"日期":a,"飯店名稱":b,"類型":c,"地區":d,"評分":e,"價錢":f,"訂房網站":g,"平均價格":h,"是否比較便宜":i,"價差":j,"倍率":k,"是否有煙火":l,"是否為國定假日":m}
#     result = col.insert_one(wk)

# print("finish")


win=tk.Tk()
win.title("訂房網站查詢")
win.geometry('800x500')
win.config(bg="#CDCD9A")

scrollbar=tk.Scrollbar(win)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
text=tk.Text(win, height=30,width=80)
text.config(padx=5, pady=5,font="Helvetic 18")
text.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=text.yview)


optionList = ['類型','飯店','整間房屋/公寓','民宿','旅館','青年旅館']
var = tk.StringVar()
var.set(optionList[0])
myoptionmenu = tk.OptionMenu(win, var, *optionList)
myoptionmenu.pack(pady=5)


optionList = ['活動日期','04/25 星期一','04/28 星期四','05/02 星期一','05/05 星期四','05/09 星期一','05/12 星期四','05/16 星期一',
            '05/19 星期四','05/21 星期六','05/23 星期一','05/26 星期四','05/28 星期六','05/30 星期一','06/02 星期四','06/06 星期一',
            '06/09 星期四','06/11 星期六','06/13 星期一','06/16 星期四','06/20 星期一','06/23 星期四','06/25 星期六','06/27 星期一','06/30 星期四']
var1 = tk.StringVar()
var1.set(optionList[0])
myoptionmenu = tk.OptionMenu(win, var1, *optionList)
myoptionmenu.pack(pady=5)


optionList = ['排序','"價錢排序"','從小到大','從大到小','---------','"評分排序"','從高到低','從低到高','---------','"折價排序"','從多到少','從少到多',]
var2 = tk.StringVar()
var2.set(optionList[0])
myoptionmenu = tk.OptionMenu(win, var2, *optionList)
myoptionmenu.pack(pady=5)


en=tk.Entry()
en.pack(pady=5)


def roomlist():
    t = int(en.get())
    data = {'價錢': {'$lt': t},'類型':var.get(),'日期':var1.get()}
    mydata = col.find(data)
    if var2.get() == "從小到大":
        mydata = mydata.sort("價錢",pymongo.ASCENDING)
        for result2 in mydata:
            strlist = []
            x = result2["飯店名稱"]
            f = result2["價錢"]
            o = result2["類型"]
            g = result2["訂房網站"]
            h = result2["平均價格"]
            i = result2["是否比較便宜"]
            j = result2["價差"]
            k = result2["倍率"]
            d = result2["地區"]
            p = result2["日期"]
            u = result2["評分"]
            strlist.append(f"類型:{o}, 名稱:{x}, 地區:{d}, 訂房網站:{g}, 評分:{u}\n價錢:${f}, 平均價格:${h}, 價差:${j}, 倍率:{k}\n")
            text.insert("insert",'\n'.join(strlist))
            text.insert("insert","\n")
            text.pack(fill=tk.Y)
            print(strlist)

    elif var2.get() == "從大到小":
        mydata = mydata.sort("價錢",pymongo.DESCENDING)
        for result2 in mydata:
            strlist = []
            x = result2["飯店名稱"]
            f = result2["價錢"]
            o = result2["類型"]
            g = result2["訂房網站"]
            h = result2["平均價格"]
            i = result2["是否比較便宜"]
            j = result2["價差"]
            k = result2["倍率"]
            p = result2["日期"]
            d = result2["地區"]
            u = result2["評分"]
            strlist.append(f"類型:{o}, 名稱:{x}, 地區:{d}, 訂房網站:{g}, 評分:{u}\n價錢:${f}, 平均價格:${h}, 價差:${j}, 倍率:{k}\n")
            text.insert("insert",'\n'.join(strlist))
            text.insert("insert","\n")
            text.pack(fill=tk.Y)

    elif var2.get() == "從高到低":
        mydata = mydata.sort("評分",pymongo.DESCENDING)
        for result2 in mydata:
            strlist = []
            x = result2["飯店名稱"]
            f = result2["價錢"]
            o = result2["類型"]
            g = result2["訂房網站"]
            h = result2["平均價格"]
            i = result2["是否比較便宜"]
            j = result2["價差"]
            k = result2["倍率"]
            d = result2["地區"]
            p = result2["日期"]
            u = result2["評分"]
            strlist.append(f"類型:{o}, 名稱:{x}, 地區:{d}, 訂房網站:{g}, 評分:{u}\n價錢:${f}, 平均價格:${h}, 價差:${j}, 倍率:{k}\n")
            text.insert("insert",'\n'.join(strlist))
            text.insert("insert","\n")
            text.pack(fill=tk.Y)

    elif var2.get() == "從低到高":
        mydata = mydata.sort("評分",pymongo.ASCENDING)
        for result2 in mydata:
            strlist = []
            x = result2["飯店名稱"]
            f = result2["價錢"]
            o = result2["類型"]
            g = result2["訂房網站"]
            h = result2["平均價格"]
            i = result2["是否比較便宜"]
            j = result2["價差"]
            k = result2["倍率"]
            p = result2["日期"]
            d = result2["地區"]
            u = result2["評分"]
            strlist.append(f"類型:{o}, 名稱:{x}, 地區:{d}, 訂房網站:{g}, 評分:{u}\n價錢:${f}, 平均價格:${h}, 價差:${j}, 倍率:{k}\n")
            text.insert("insert",'\n'.join(strlist))
            text.insert("insert","\n")
            text.pack(fill=tk.Y)

    elif var2.get() == "從多到少":
        mydata = mydata.sort("價差",pymongo.ASCENDING)
        for result2 in mydata:
            strlist = []
            x = result2["飯店名稱"]
            f = result2["價錢"]
            o = result2["類型"]
            g = result2["訂房網站"]
            h = result2["平均價格"]
            d = result2["地區"]
            j = result2["價差"]
            k = result2["倍率"]
            p = result2["日期"]
            u = result2["評分"]
            strlist.append(f"類型:{o}, 名稱:{x}, 地區:{d}, 訂房網站:{g}, 評分:{u}\n價錢:${f}, 平均價格:${h}, 價差:${j}, 倍率:{k}\n")
            text.insert("insert",'\n'.join(strlist))
            text.insert("insert","\n")
            text.pack(fill=tk.Y)

    elif var2.get() == "從少到多":
        mydata = mydata.sort("價差",pymongo.DESCENDING)
        for result2 in mydata:
            strlist = []
            x = result2["飯店名稱"]
            f = result2["價錢"]
            o = result2["類型"]
            d = result2["地區"]
            g = result2["訂房網站"]
            h = result2["平均價格"]
            i = result2["是否比較便宜"]
            j = result2["價差"]
            k = result2["倍率"]
            p = result2["日期"]
            u = result2["評分"]
            strlist.append(f"類型:{o}, 名稱:{x}, 地區:{d}, 訂房網站:{g}, 評分:{u}\n價錢:${f}, 平均價格:${h}, 價差:${j}, 倍率:{k}\n")
            text.insert("insert",'\n'.join(strlist))
            text.insert("insert","\n")
            text.pack(fill=tk.Y)



def clear():
    text.delete()


btn=tk.Button(text='搜尋')
btn.config(bg='#00FFFF',fg='#00008B',command=roomlist)
btn.pack(pady=5)


btn1=tk.Button(text='清除',command = lambda:text.delete(1.0,tk.END))
btn1.config(bg='#00FFFF',fg='#00008B')
btn1.pack(pady=5)


win.mainloop() 