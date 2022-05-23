import requests
from bs4 import BeautifulSoup
import datetime
import random
import time
import openpyxl

wb = openpyxl.Workbook() #開啟新的工作簿
ws = wb.active           #開啟新的工作表

ws["A1"] = "政府單位"
ws["B1"] = "採購案"
ws["C1"] = "網址"
ws["D1"] = "起始日"
ws["E1"] = "截止日"
ws["F1"] = "採購金額"

data = {
    "tmpQuerySentence":None,
    "timeRange":"111/1/1-111/12/31",
    "chkcode":None,
    "querySentence":"教育訓練",
    "tenderStatusType":"招標",
    "sortCol":"TENDER_NOTICE_DATE",
    "timeRangeTemp":"111/1/1-111/12/31",
    "sym":"on",
    "itemPerPage":"50"
}

res = requests.post("http://web.pcc.gov.tw/prkms/prms-searchBulletionClient.do?root=tps",data)
soup = BeautifulSoup(res.text)

case_num=0
while case_num<50:
    case_num+=1
    if soup.find_all('td',headers="addThId_i")[0].select('tr')[0].select('tr')[case_num].select('td')[4].string == str(int(str(datetime.date.today()).replace('-','/')[:4])-1911)+str(datetime.date.today()).replace('-','/')[4:]:
        res2=requests.get('http://web.pcc.gov.tw'+soup.find_all('td',headers="addThId_i")[0].select('tr')[0].select('tr')[case_num].select('td')[3].a['href'])
        soup2=BeautifulSoup(res2.text)
        #print(soup.find_all('td',headers="addThId_i")[0].select('tr')[0].select('tr')[1])
        a = soup.find_all('td',headers="addThId_i")[0].select('tr')[0].select('tr')[case_num].select('td')[2].string
        b = soup.find_all('td',headers="addThId_i")[0].select('tr')[0].select('tr')[case_num].select('td')[3].div.string
        c = 'http://web.pcc.gov.tw'+soup.find_all('td',headers="addThId_i")[0].select('tr')[0].select('tr')[case_num].select('td')[3].a['href']
        d = soup.find_all('td',headers="addThId_i")[0].select('tr')[0].select('tr')[case_num].select('td')[4].string
        e = soup.find_all('td',headers="addThId_i")[0].select('tr')[0].select('tr')[case_num].select('td')[6].string
        for i in range(45,60):
            if soup2.select('tr')[i].th.text.strip() == "預算金額":
                f = soup2.select('tr')[i].td.text.strip()
                break
            # for i in range(10):
            #     if soup2.select('tr')[57+i].td.text.strip() == "否" or soup2.select('tr')[57+i].td.text.strip() == "是":
            #         continue
                # print(soup2.select('tr')[54+i].td.text.strip())
                # else:
                #     print(soup2.select('tr')[57+i].td.text.strip())
                #     break
                # else:
                #     print(soup2.select('tr')[57].td.text.strip())
        ws.append([a,b,c,d,e,f])
        wb.save("採購案.xlsx")
        print(case_num)
        time.sleep(random.randint(8,20))

