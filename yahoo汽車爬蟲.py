import requests
from bs4 import BeautifulSoup
import openpyxl as op
import time

# wb =op.Workbook("汽車搜尋1.xlsx")
wb = op.load_workbook("汽車搜尋1.xlsx")
ws = wb.active

ws["A1"] = "名稱"
ws["B1"] = "價格"
ws["C1"] = "動力"
ws["D1"] = "網址"

res=requests.get("https://autos.yahoo.com.tw/car-search?price_range=150-200")
soup=BeautifulSoup(res.text)
car=soup.find_all("div",class_="model-wrapper")
price=["0.01-60","60-100","100-150","150-200","200-250","250-400","400"]
for k in range(len(price)):
    for j in range(1,10):
        res=requests.get("https://autos.yahoo.com.tw/car-search?price_range="+price[k]+"&p="+str(j))
        soup=BeautifulSoup(res.text)
        car=soup.find_all("div",class_="model-wrapper")
        if car==[]:
            break
        else:
            for i in range(len(car)):
                a = (car[i].find_all("span")[0].text.replace("\n","")) #名稱
                b = (car[i].find_all("span")[1].text.replace(" ","").replace("\n","")) #價格
                c = (car[i].find_all("li",class_="model-sub")[0].ul.find_all("li")[-1].text) #動力
                d = (car[i].ul.a["href"]) #網址
                ws.append([a,b,c,d])
                print("------------------------------------")
            print(f"{k}第{j}頁")
            time.sleep(5)

ws = wb.save("汽車搜尋1.xlsx")

# for x in range(1,4):
#     res=requests.get(f"https://autos.yahoo.com.tw/car-search?price_range=60-100&p={x}")
#     soup=BeautifulSoup(res.text)
#     for car in soup.find_all("div",class_="model-wrapper"):
#         print(car.find_all("span")[0].text.replace(" ","").replace("\n",""))
#         print(car.find_all("span")[1].text.replace(" ","").replace("\n",""),"萬")
#         print(car.find_all("li",class_="model-sub")[0].ul.find_all("li")[-1].text)
#         print(soup.find_all("div",class_="model-wrapper")[0].ul.a["href"])
#     print(f"--------------第{x}頁-----------------")




