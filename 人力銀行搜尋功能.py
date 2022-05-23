import requests 
import openpyxl
from bs4 import BeautifulSoup
import tkinter as tk
from datetime import datetime
import time
import random
wb=openpyxl.Workbook()
ws=wb.active
date=datetime.now().strftime('%Y-%m-%d')
win=tk.Tk()
win.title("1111職缺查詢")
win.geometry('300x300')
win.config(bg="#CDCD9A")

lb=tk.Label(text="職缺查詢")
lb.config(bg="#C6A300",fg="#4F4F4F")
lb.pack()

en1=tk.Entry()
en1.pack()


ws['B1']='職缺連結'
ws['C1']='公司名稱'
ws['D1']='工作地區'
ws['E1']='薪資待遇'
ws['F1']='給薪方式'
ws['G1']='薪資下限'
ws['H1']='薪資上限'
lb1 = ""


    #讀取搜索頁(以大數據為關鍵字)
def f():
    t=en1.get()
    res=requests.get('https://www.1111.com.tw/search/job?ks='+t)
    soup=BeautifulSoup(res.text)

    #讀取page頁的while迴圈
    page=0
    while page<2:
        page+=1
        #讀取page頁，覆蓋res→soup
        res=requests.get('https://www.1111.com.tw/search/job?ks='+t+'&page='+str(page))
        soup=BeautifulSoup(res.text)
        
        i =page
        i =tk.Label()
        i.config(text=page,bg="#C6A300",fg="#4F4F4F")
        i.pack()

        #讀取單一搜索頁的各職缺(一頁20個)
        for job in soup.find_all('div',class_="srh-body__result-item srh-body__result-item--loaded")[0].find_all('div',class_="job_item_info"):
            #二階爬蟲：以各職缺連結進一步抓取職缺頁內的資料存在res2→soup2
            #res2=requests.get(job.a['href'])
            #soup2=BeautifulSoup(res2.text)
            #印出各職缺的資料
            a=job.a.h5.text#職缺名稱
            b=job.a['href']#職缺連結
            c=job.h6.text#公司名稱
            d=job.select('div')[2].div.a.text#公司地區
            e=job.select('div')[2].div.div.text#薪資待遇
            f=job.select('div')[2].div.div.text[0:2]#給薪方式

            price=''
            for word in e:
                if word=='0' or word=='1' or word=='2' or word=='3' or word=='4' or word=='5' or word=='6' or word=='7' or word=='8' or word=='9' or word=='.' or word=='~' or word=='萬':
                    price+=word

            low_price=''
            high_price=''
            if '~' in price:
                low_price=price[0:price.find('~')]
                high_price=price[price.find('~')+1:len(price)]
            else:
                low_price=price
                high_price=price

            if '萬' in low_price:
                low_price=low_price[0:low_price.find('萬')]
                low_price=float(low_price)*10000
            else:
                low_price=float(low_price)

            if '萬' in high_price:
                high_price=high_price[0:high_price.find('萬')]
                high_price=float(high_price)*10000
            else:
                high_price=float(high_price)        

            ws.append([a,b,c,d,e,f,low_price,high_price])

            #print(soup2.find_all('div',class_="contacter_info")[0].p.text)#聯絡人
            #time.sleep(random.randint(3,10))

        print('已讀取',page,'頁')
        print('-------------------------------------------')
        time.sleep(random.randint(2,5))
        wb.save('%s%s.xlsx'%(t,date))
    lb2=tk.Label()
    lb2.config(text = "Finish!",bg="#C6A300",fg="#4F4F4F")
    lb2.pack()

btn=tk.Button(text="開始搜尋")
btn.config(command=f)
btn.pack()
win.mainloop() 